import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 定义文件路径
input_file = '/Users/ml25qs01/trae-workspace/千瓜品牌数据抓取/千瓜数据_古茗_抖音_2026-03-01_to_2026-03-31.xlsx'
output_file = '/Users/ml25qs01/trae-workspace/千瓜品牌数据抓取/千瓜数据_古茗_抖音_2026-03-01_to_2026-03-31（排重）.xlsx'

# 读取Excel文件
wb = load_workbook(input_file)
ws = wb.active

# 提取271-290行C列的标题（注意：Excel行号从1开始）
target_titles = []
for row in range(271, 291):  # 271到290行
    cell_value = ws[f'C{row}'].value
    if cell_value:
        target_titles.append(str(cell_value).strip())

# 定义低饱和度黄色填充
yellow_fill = PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid')

# 检查前268行数据中是否有与目标标题重复的
marked_count = 0
for row in range(1, 269):  # 前268行
    cell_value = ws[f'C{row}'].value
    if cell_value:
        cell_str = str(cell_value).strip()
        if cell_str in target_titles:
            ws[f'C{row}'].fill = yellow_fill
            marked_count += 1

# 保存为新文件
wb.save(output_file)

print(f"处理完成！")
print(f"提取的目标标题数量: {len(target_titles)}")
print(f"标记的重复项数量: {marked_count}")
print(f"保存路径: {output_file}")
